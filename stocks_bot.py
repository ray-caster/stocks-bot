# stocks_bot.py
import os
import io
import yfinance
import datetime
import matplotlib.pyplot as plt
import matplotlib.dates
import discord
from dotenv import load_dotenv
from discord import app_commands
import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles import PatternFill,Font
load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')
intents = discord.Intents.default()
intents.message_content = True
client = discord.Client(intents=intents)
tree = app_commands.CommandTree(client)

def del_com(input):
    splitted=str(input).split(",")
    return "".join(splitted)

@tree.command(name = "graph", description = "Graph a value of a stock code. ( ie. Close, Open, High, Low, Volume )", guild=discord.Object(id=1164972452988854396))
async def plot_closing(interaction, code: str, key: str):
    data_stream = io.BytesIO()
    date_start=(datetime.datetime.today()-datetime.timedelta(days=140))
    date_end=datetime.datetime.today()
    data = yfinance.download(code+".JK", start=date_start.strftime("%Y-%m-%d"), end=date_end.strftime("%Y-%m-%d"))
    plot_data=[[],[]]
    for val in data.index:
        plot_data[0].append(val)
    for val in data[key]:
        plot_data[1].append(val)
    plt.plot(plot_data[0],plot_data[1],marker='o')
    myFmt = matplotlib.dates.DateFormatter("%d-%b")
    plt.gca().xaxis.set_major_formatter(myFmt)
    plt.xticks(rotation=90, fontweight='light',  fontsize='x-small',)
    plt.xlabel('Date')
    plt.ylabel(code+' '+key+' Values')
    plt.grid(True)
    plt.savefig(data_stream, format='png', bbox_inches="tight", dpi = 100)
    plt.close()
    data_stream.seek(0)
    chart = discord.File(data_stream,filename=(key+".png"))
    embed=discord.Embed()
    embed.set_image(
    url=("attachment://"+key+".png")
    )
    await interaction.response.send_message(embed=embed,file=chart)
@tree.command(name = "get", description = "Get the data from the past 100 days of a stock.", guild=discord.Object(id=1164972452988854396))
async def get(interaction, code: str):
    
    workbook = openpyxl.Workbook()
    data = yfinance.download(code+".JK", start=(datetime.datetime.today()-datetime.timedelta(days=140)).strftime("%Y-%m-%d"), end=datetime.datetime.today().strftime("%Y-%m-%d"))
    worksheet=workbook["Sheet"]
    worksheet.column_dimensions['A'].width = 15
    worksheet["A1"]="Date"
    worksheet["B1"]="Open"
    worksheet["C1"]="High"
    worksheet["D1"]="Low"
    worksheet["E1"]="Close*"
    worksheet["F1"]="Volume"
    max_cl=0
    for i in range (1,7):
        worksheet.cell(1,i).fill=PatternFill(start_color="70ad47", fill_type="solid")
        worksheet.cell(1,i).font=Font(color="FFFFFF",bold=True)
    
    i=0
    maxd=data.index.size+1
    
    for val in data.index:
        cell=worksheet.cell(maxd-i,1)
        cell.value=val.strftime("%Y-%m-%d")
        if((maxd-i)%2==0):
            cell.fill=PatternFill(start_color="e2efda", fill_type="solid")
        i+=1
            
    i=0
    for val in data['Open']:
        cell=worksheet.cell(maxd-i,2)
        cell.value=val
        if((maxd-i)%2==0):
            cell.fill=PatternFill(start_color="e2efda", fill_type="solid")
        i+=1
    i=0
    for val in data['High']:
        cell=worksheet.cell(maxd-i,3)
        cell.value=val
        if((maxd-i)%2==0):
            cell.fill=PatternFill(start_color="e2efda", fill_type="solid")
        i+=1
    i=0
    for val in data['Low']:
        cell=worksheet.cell(maxd-i,4)
        cell.value=val
        if((maxd-i)%2==0):
            cell.fill=PatternFill(start_color="e2efda", fill_type="solid")
        i+=1
    i=0
    for val in data['Close']:
        cell=worksheet.cell(maxd-i,5)
        cell.value=val
        if((maxd-i)%2==0):
            cell.fill=PatternFill(start_color="e2efda", fill_type="solid")
        i+=1
        max_cl=max(max_cl,val)
    i=0
    for val in data['Volume']:
        cell=worksheet.cell(maxd-i,6)
        cell.value=val
        if((maxd-i)%2==0):
            cell.fill=PatternFill(start_color="e2efda", fill_type="solid")
        i+=1
    worksheet["H1"]="CH"
    worksheet["I1"]="CL"
    worksheet["J1"]="CC"
    worksheet["K1"]="Avg Harian"
    worksheet["L1"]="MA5"
    worksheet["M1"]="Op=Low"
    worksheet["N1"]="Op=High"
    worksheet["O1"]="Prank"
    worksheet["P1"]="JJS OpLo"
    worksheet["Q1"]="WR JJSOL"
    worksheet["R1"]="ProsentasePrank"
    worksheet.column_dimensions['R'].width = 15
    worksheet["S1"]="OpLo9"
    worksheet["T1"]="WR OpLo9"
    winsjjs=0
    losesjjs=0
    winsoplo9=0
    losesoplo9=0
    pranks=0
    for i in range(2,worksheet.max_row):
        worksheet.cell(i,8).value=(float)(del_com(worksheet.cell(i,3).value))/(float)(del_com(worksheet.cell(i+1,5).value))-1
        if(worksheet.cell(i,8).value>=0.02):
            worksheet.cell(i,8).fill=PatternFill(start_color="008000",fill_type="solid")
        worksheet.cell(i,8).number_format=numbers.FORMAT_PERCENTAGE_00
        worksheet.cell(i,9).value=(float)(del_com(worksheet.cell(i,4).value))/(float)(del_com(worksheet.cell(i+1,5).value))-1
        if(worksheet.cell(i,9).value<=-0.03):
            worksheet.cell(i,9).fill=PatternFill(start_color="800000",fill_type="solid")
            worksheet.cell(i,9).font=Font(color="FFFFFF")
        worksheet.cell(i,9).number_format=numbers.FORMAT_PERCENTAGE_00
        worksheet.cell(i,10).value=(float)(del_com(worksheet.cell(i,5).value))/(float)(del_com(worksheet.cell(i+1,5).value))-1
        if(worksheet.cell(i,10).value<=-0.03):
            worksheet.cell(i,10).fill=PatternFill(start_color="800000",fill_type="solid")
            worksheet.cell(i,10).font=Font(color="FFFFFF")
        worksheet.cell(i,10).number_format=numbers.FORMAT_PERCENTAGE_00
        worksheet.cell(i,11).value=((float)(del_com(worksheet.cell(i,2).value))+(float)(del_com(worksheet.cell(i,3).value))+(float)(del_com(worksheet.cell(i,4).value))+(float)(del_com(worksheet.cell(i,5).value)))/4
        if(i-4>=2):
            worksheet.cell(i-4,12).value=((float)(del_com(worksheet.cell(i,11).value))+(float)(del_com(worksheet.cell(i-1,11).value))+(float)(del_com(worksheet.cell(i-2,11).value))+(float)(del_com(worksheet.cell(i-3,11).value))+(float)(del_com(worksheet.cell(i-4,11).value)))/5
        if(del_com(worksheet.cell(i,4).value)==del_com(worksheet.cell(i,2).value)):
            worksheet.cell(i,13).value="YES"
        else:
            worksheet.cell(i,13).value="---"
        if(del_com(worksheet.cell(i,3).value)==del_com(worksheet.cell(i,2).value)):
            worksheet.cell(i,14).value="YES"
        else:
            worksheet.cell(i,14).value="---"
        if(del_com(worksheet.cell(i,3).value)==del_com(worksheet.cell(i,2).value) and worksheet.cell(i,8).value>=0.02):
            worksheet.cell(i,15).value="YES"
            pranks+=1
        else:
            worksheet.cell(i,15).value="---"
        if(del_com(worksheet.cell(i,4).value)==del_com(worksheet.cell(i,2).value)  and i>2):
            if(worksheet.cell(i-1,8).value>=0.02):
                worksheet.cell(i,16).value="WIN"
                winsjjs+=1
            else:
                worksheet.cell(i,16).value="LOSE"
                losesjjs+=1
        else:
            worksheet.cell(i,16).value="---"
        if(del_com(worksheet.cell(i,4).value)==del_com(worksheet.cell(i,2).value)  and i>2):
            if(((float)(del_com(worksheet.cell(i,3).value))/(float)(del_com(worksheet.cell(i,2).value)))>=1.03):
                worksheet.cell(i,19).value="WIN"
                winsoplo9+=1
            else:
                worksheet.cell(i,19).value="LOSE"
                losesoplo9+=1
        else:
            worksheet.cell(i,19).value="---"
    if(winsjjs+losesjjs>0):
        worksheet.cell(2,17).value=winsjjs/(winsjjs+losesjjs)
    else:
        worksheet.cell(2,17).value=0
    worksheet.cell(2,17).number_format=numbers.FORMAT_PERCENTAGE_00
    worksheet.cell(2,18).value=pranks/(worksheet.max_row-1)
    worksheet.cell(2,18).number_format=numbers.FORMAT_PERCENTAGE_00
    if(winsoplo9+losesoplo9>0):
        worksheet.cell(2,20).value=winsoplo9/(winsoplo9+losesoplo9)
    else:
        worksheet.cell(2,20).value=0
    worksheet.cell(2,20).number_format=numbers.FORMAT_PERCENTAGE_00
    data_stream = io.BytesIO()
    workbook.save(data_stream)
    data_stream.seek(0)
    img = discord.File(data_stream,filename=(code+".xlsx"))
    embed=discord.Embed()
    embed.set_image(
    url=("attachment://"+code+".xlsx")
    )
    await interaction.response.send_message(file=img)
@client.event
async def on_ready():
    await tree.sync(guild=discord.Object(id=1164972452988854396))
    print("Ready!")

client.run(TOKEN)

